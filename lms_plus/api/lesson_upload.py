import frappe
import json
import io
import base64
import random
import string
import re
from frappe import _


def _rand_id(n=10):
    return "".join(random.choices(string.ascii_letters + string.digits, k=n))


def _extract_yt_id(url):
    for pat in [
        r"youtu\.be/([a-zA-Z0-9_-]{11})",
        r"youtube\.com/watch\?v=([a-zA-Z0-9_-]{11})",
        r"youtube\.com/embed/([a-zA-Z0-9_-]{11})",
    ]:
        m = re.search(pat, url)
        if m:
            return m.group(1)
    return None


def _quiz_exists(quiz_id):
    """Check if a quiz exists in the LMS before adding it to content."""
    if not quiz_id:
        return False
    return bool(frappe.db.exists("LMS Quiz", quiz_id))


def _build_content(youtube_url=None, content_text=None, quiz_id=None):
    """Build EditorJS JSON from individual fields."""
    import time as _time
    blocks = []

    if youtube_url:
        vid = _extract_yt_id(youtube_url)
        if vid:
            blocks.append({
                "id": _rand_id(),
                "type": "embed",
                "data": {
                    "service": "youtube",
                    "source": youtube_url,
                    "embed": vid,
                    "caption": "",
                },
            })

    if content_text:
        blocks.append({
            "id": _rand_id(),
            "type": "markdown",
            "data": {"text": str(content_text)},
        })

    # Only add quiz block if the quiz actually exists in the LMS
    if quiz_id and _quiz_exists(quiz_id):
        blocks.append({
            "id": _rand_id(),
            "type": "quiz",
            "data": {"quiz": str(quiz_id)},
        })
    elif quiz_id:
        frappe.logger().warning(
            f"LMS Plus lesson upload: Quiz '{quiz_id}' not found — skipping quiz block."
        )

    if not blocks:
        blocks.append({
            "id": _rand_id(),
            "type": "markdown",
            "data": {"text": ""},
        })

    return json.dumps({
        "time": int(_time.time() * 1000),
        "blocks": blocks,
        "version": "2.29.0",
    })


@frappe.whitelist()
def download_template():
    """Returns a base64-encoded Excel template file for lesson bulk upload."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lessons"

    headers = [
        "Chapter Title",
        "Lesson Title",
        "YouTube URL",
        "Content (Text / Markdown)",
        "Quiz ID",
        "Include in Preview (Yes/No)",
    ]

    fill   = PatternFill(start_color="1A7A6E", end_color="1A7A6E", fill_type="solid")
    font   = Font(color="FFFFFF", bold=True)
    centre = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = centre

    samples = [
        ["Introduction to Python", "What is Python?",
         "https://www.youtube.com/watch?v=example1",
         "Python is a popular programming language.", "", "Yes"],
        ["Introduction to Python", "Setting Up Python",
         "https://www.youtube.com/watch?v=example2",
         "How to install Python on your computer.", "python-setup-quiz", "No"],
        ["Advanced Topics", "Functions in Python", "",
         "Functions are reusable blocks of code.", "", "No"],
    ]
    for s in samples:
        ws.append(s)

    for col, w in enumerate([22, 28, 48, 42, 22, 28], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws2 = wb.create_sheet("Instructions")
    for row in [
        ["LESSON BULK UPLOAD — HOW TO USE"],
        [""],
        ["Column", "Description", "Required"],
        ["Chapter Title", "Chapter name. Created automatically if it does not exist.", "Yes"],
        ["Lesson Title", "Name of the lesson.", "Yes"],
        ["YouTube URL", "Full YouTube URL (https://www.youtube.com/watch?v=...)", "No"],
        ["Content", "Plain text or Markdown content for the lesson body.", "No"],
        ["Quiz ID", "Quiz slug to attach to this lesson (must already exist in LMS).", "No"],
        ["Include in Preview", "Type Yes to allow non-enrolled users to preview.", "No"],
        [""],
        ["RULES"],
        ["Rows with empty Chapter Title or Lesson Title are skipped."],
        ["Chapters and lessons are created in the order they appear in the file."],
        ["Lessons with the same title in the same chapter are skipped (no duplicates)."],
    ]:
        ws2.append(row)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 65
    ws2.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = base64.b64encode(buf.read()).decode("utf-8")
    return {"file_content": encoded, "filename": "lesson_bulk_upload_template.xlsx"}


@frappe.whitelist()
def process_lesson_upload(course: str, file_url: str):
    """
    Reads the uploaded Excel and enqueues lesson creation as a background job.
    """
    import openpyxl

    if not frappe.db.exists("LMS Course", course):
        frappe.throw(_("Course {0} not found.").format(course))

    site_path = frappe.get_site_path()
    if file_url.startswith("/files/"):
        file_path = site_path + "/public" + file_url
    elif file_url.startswith("/private/files/"):
        file_path = site_path + file_url
    else:
        frappe.throw(_("Invalid file URL."))

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Lessons"]

    HEADER_SKIP = {"chapter title", "lesson title"}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch_title  = str(row[0] or "").strip()
        les_title = str(row[1] or "").strip()
        if not ch_title or not les_title:
            continue
        # Skip rows that look like header labels
        if ch_title.lower() in HEADER_SKIP or les_title.lower() in HEADER_SKIP:
            continue
        rows.append({
            "chapter_title":    ch_title,
            "lesson_title":     les_title,
            "youtube_url":      str(row[2] or "").strip(),
            "content_text":     str(row[3] or "").strip(),
            "quiz_id":          str(row[4] or "").strip(),
            "include_in_preview": str(row[5] or "").strip().lower() in ["yes", "y", "1", "true"],
        })

    if not rows:
        frappe.throw(_(
            "No valid rows found in the uploaded file. "
            "Check that the sheet is named Lessons and rows are not empty."
        ))

    frappe.enqueue(
        "lms_plus.api.lesson_upload._create_lessons_job",
        course=course,
        rows=rows,
        queue="long",
        timeout=3600,
    )

    return {
        "status": "enqueued",
        "count": len(rows),
        "message": _("{0} lessons queued for creation in the background.").format(len(rows)),
    }


def _create_lessons_job(course: str, rows: list):
    """
    Background job — creates chapters and lessons from Excel rows.
    """
    course_doc = frappe.get_doc("LMS Course", course)

    # Map existing chapter titles (lower) -> chapter name
    existing_chapters = {}
    for ref in course_doc.chapters:
        ch = frappe.get_cached_doc("Course Chapter", ref.chapter)
        existing_chapters[ch.title.strip().lower()] = ch.name

    # Group rows by chapter preserving file order
    chapter_order = []
    chapter_rows  = {}
    for row in rows:
        key = row["chapter_title"].strip().lower()
        if key not in chapter_rows:
            chapter_rows[key] = {"title": row["chapter_title"], "lessons": []}
            chapter_order.append(key)
        chapter_rows[key]["lessons"].append(row)

    created_ch, created_les, skipped = 0, 0, 0

    for ch_key in chapter_order:
        ch_data  = chapter_rows[ch_key]
        ch_title = ch_data["title"]

        if ch_key in existing_chapters:
            chapter_doc = frappe.get_doc("Course Chapter", existing_chapters[ch_key])
        else:
            chapter_doc = frappe.get_doc({
                "doctype":      "Course Chapter",
                "title":        ch_title,
                "course":       course,
                "course_title": course_doc.title,
            })
            chapter_doc.insert(ignore_permissions=True)
            existing_chapters[ch_key] = chapter_doc.name

            course_doc.reload()
            course_doc.append("chapters", {"chapter": chapter_doc.name})
            course_doc.save(ignore_permissions=True)
            created_ch += 1

        # Existing lesson titles in this chapter
        existing_titles = set()
        for lr in chapter_doc.lessons:
            t = frappe.db.get_value("Course Lesson", lr.lesson, "title")
            if t:
                existing_titles.add(t.strip().lower())

        for lesson_row in ch_data["lessons"]:
            les_title = lesson_row["lesson_title"].strip()

            if les_title.lower() in existing_titles:
                skipped += 1
                continue

            content = _build_content(
                youtube_url  = lesson_row["youtube_url"]  or None,
                content_text = lesson_row["content_text"] or None,
                quiz_id      = lesson_row["quiz_id"]      or None,
            )

            les_doc = frappe.get_doc({
                "doctype":           "Course Lesson",
                "title":             les_title,
                "chapter":           chapter_doc.name,
                "course":            course,
                "content":           content,
                "include_in_preview": 1 if lesson_row["include_in_preview"] else 0,
            })
            les_doc.insert(ignore_permissions=True)

            chapter_doc.reload()
            chapter_doc.append("lessons", {"lesson": les_doc.name})
            chapter_doc.save(ignore_permissions=True)

            existing_titles.add(les_title.lower())
            created_les += 1

        frappe.db.commit()

    frappe.logger().info(
        f"LMS Plus lesson upload '{course}': "
        f"{created_ch} chapters, {created_les} lessons created, {skipped} skipped."
    )
