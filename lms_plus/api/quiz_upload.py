import frappe
import json
import io
import base64
import re
from frappe import _


def _slugify(text):
    """Convert text to a URL-friendly slug."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s-]+", "-", text)
    return text[:60]


@frappe.whitelist()
def download_quiz_template():
    """Returns a base64-encoded Excel template for bulk quiz upload."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()

    teal_fill = PatternFill(start_color="1A7A6E", end_color="1A7A6E", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    norm = Font(size=10)
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # ── Sheet 1: Quiz Settings ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Quiz Settings"
    ws1.freeze_panes = "A2"

    s1_headers = [
        "Quiz Title",
        "Quiz Slug (URL ID)",
        "Passing % (0-100)",
        "Max Attempts",
        "Duration (minutes)",
        "Shuffle Questions (Yes/No)",
        "Limit Questions To",
        "Negative Marking (Yes/No)",
        "Marks to Cut",
    ]
    s1_widths = [28, 28, 18, 16, 20, 24, 20, 24, 16]

    for col, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = teal_fill
        cell.font = white_font
        cell.alignment = centre

    for col, w in enumerate(s1_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = w

    ws1.row_dimensions[1].height = 28

    s1_samples = [
        ["Python Basics Quiz", "python-basics-quiz", 60, 3, 15, "Yes", 5, "No", 0],
        ["Data Types Quiz",    "data-types-quiz",    70, 2, 10, "Yes", 3, "Yes", 1],
    ]
    for i, row in enumerate(s1_samples, 2):
        ws1.row_dimensions[i].height = 20
        fill = gray_fill if i % 2 == 0 else None
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = norm
            cell.alignment = left
            if fill:
                cell.fill = fill

    # ── Sheet 2: Questions ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Questions")
    ws2.freeze_panes = "A2"

    s2_headers = [
        "Quiz Slug",
        "Question",
        "Type (Choices / User Input)",
        "Multiple Correct (Yes/No)",
        "Option 1", "Correct 1 (Yes/No)",
        "Option 2", "Correct 2 (Yes/No)",
        "Option 3", "Correct 3 (Yes/No)",
        "Option 4", "Correct 4 (Yes/No)",
        "Possible Answer (User Input only)",
        "Marks",
    ]
    s2_widths = [26, 50, 26, 22, 22, 20, 22, 20, 22, 20, 22, 20, 32, 10]

    for col, h in enumerate(s2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = teal_fill
        cell.font = white_font
        cell.alignment = centre

    for col, w in enumerate(s2_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = w

    ws2.row_dimensions[1].height = 28

    s2_samples = [
        ["python-basics-quiz", "What does Python stand for?", "Choices", "No",
         "A programming language", "Yes", "A snake", "No",
         "A software framework", "No", "A database", "No", "", 1],
        ["python-basics-quiz", "Which of these are Python data types?", "Choices", "Yes",
         "int", "Yes", "float", "Yes",
         "char", "No", "boolean", "Yes", "", 2],
        ["python-basics-quiz", "What keyword is used to define a function in Python?", "User Input", "No",
         "", "", "", "", "", "", "", "", "def", 1],
        ["data-types-quiz", "Which data type stores True or False?", "Choices", "No",
         "bool", "Yes", "int", "No",
         "str", "No", "float", "No", "", 1],
    ]
    for i, row in enumerate(s2_samples, 2):
        ws2.row_dimensions[i].height = 20
        fill = gray_fill if i % 2 == 0 else None
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = norm
            cell.alignment = left
            if fill:
                cell.fill = fill

    # ── Sheet 3: Instructions ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Instructions")
    rows = [
        ("BULK QUIZ UPLOAD — INSTRUCTIONS", True, "1A7A6E", "FFFFFF"),
        ("", False, None, None),
        ("HOW IT WORKS", True, "374151", None),
        ("1. Fill the Quiz Settings sheet — one row per quiz.", False, None, None),
        ("2. Fill the Questions sheet — one row per question. Use the same Quiz Slug to link questions to a quiz.", False, None, None),
        ("3. Upload the file via the Upload Quizzes button on the course dashboard.", False, None, None),
        ("4. Quizzes and questions are created in the background. Refresh the page after 15 seconds.", False, None, None),
        ("", False, None, None),
        ("UNIQUE QUIZZES PER LEARNER", True, "374151", None),
        ("Set Shuffle Questions to Yes and set Limit Questions To (e.g. 5).", False, None, None),
        ("If you upload 20 questions and set Limit to 5, each learner gets 5 different random questions.", False, None, None),
        ("", False, None, None),
        ("RULES", True, "374151", None),
        ("Quiz Slug must be unique. If a quiz with the same slug already exists it will be skipped.", False, None, None),
        ("Questions with empty Question text are skipped.", False, None, None),
        ("Type must be exactly: Choices or User Input.", False, None, None),
        ("For User Input questions, leave all Option columns blank.", False, None, None),
        ("Correct column accepts: Yes, No, y, n, 1, 0, true, false.", False, None, None),
        ("Marks column must be a whole number (default is 1 if left blank).", False, None, None),
    ]
    for i, (text, bold, bg, fg) in enumerate(rows, 1):
        cell = ws3.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=10, color=fg or "111111")
        if bg:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 90
    for i in range(1, len(rows) + 2):
        ws3.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = base64.b64encode(buf.read()).decode("utf-8")
    return {"file_content": encoded, "filename": "quiz_bulk_upload_template.xlsx"}


@frappe.whitelist()
def process_quiz_upload(file_url: str):
    """
    Reads uploaded Excel and enqueues quiz and question creation.
    """
    import openpyxl

    site_path = frappe.get_site_path()
    if file_url.startswith("/files/"):
        file_path = site_path + "/public" + file_url
    elif file_url.startswith("/private/files/"):
        file_path = site_path + file_url
    else:
        frappe.throw(_("Invalid file URL."))

    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "Quiz Settings" not in wb.sheetnames:
        frappe.throw(_("Quiz Settings sheet not found. Please use the provided template."))
    if "Questions" not in wb.sheetnames:
        frappe.throw(_("Questions sheet not found. Please use the provided template."))

    # Parse Quiz Settings
    ws1 = wb["Quiz Settings"]
    quizzes = {}
    SETTINGS_SKIP = {"quiz title", "quiz slug (url id)"}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        title = str(row[0] or "").strip()
        slug  = str(row[1] or "").strip()
        if not title or not slug:
            continue
        if title.lower() in SETTINGS_SKIP or slug.lower() in SETTINGS_SKIP:
            continue
        quizzes[slug] = {
            "title":               title,
            "slug":                slug or _slugify(title),
            "passing_percentage":  int(row[2] or 0),
            "max_attempts":        int(row[3] or 3),
            "duration":            str(row[4] or ""),
            "shuffle_questions":   str(row[5] or "").strip().lower() in ["yes", "y", "1", "true"],
            "limit_questions_to":  int(row[6] or 0),
            "enable_negative_marking": str(row[7] or "").strip().lower() in ["yes", "y", "1", "true"],
            "marks_to_cut":        int(row[8] or 0),
            "questions":           [],
        }

    if not quizzes:
        frappe.throw(_("No valid quizzes found in the Quiz Settings sheet."))

    # Parse Questions
    ws2 = wb["Questions"]
    QUESTION_SKIP = {"quiz slug", "question"}

    def is_correct(val):
        return str(val or "").strip().lower() in ["yes", "y", "1", "true"]

    for row in ws2.iter_rows(min_row=2, values_only=True):
        slug  = str(row[0] or "").strip()
        qtext = str(row[1] or "").strip()
        if not slug or not qtext:
            continue
        if slug.lower() in QUESTION_SKIP or qtext.lower() in QUESTION_SKIP:
            continue
        if slug not in quizzes:
            continue

        qtype    = str(row[2] or "Choices").strip()
        multiple = str(row[3] or "").strip().lower() in ["yes", "y", "1", "true"]

        quizzes[slug]["questions"].append({
            "question":     qtext,
            "type":         qtype if qtype in ["Choices", "User Input"] else "Choices",
            "multiple":     multiple,
            "option_1":     str(row[4] or "").strip(),
            "correct_1":    is_correct(row[5]),
            "option_2":     str(row[6] or "").strip(),
            "correct_2":    is_correct(row[7]),
            "option_3":     str(row[8] or "").strip(),
            "correct_3":    is_correct(row[9]),
            "option_4":     str(row[10] or "").strip(),
            "correct_4":    is_correct(row[11]),
            "possibility_1": str(row[12] or "").strip(),
            "marks":        int(row[13] or 1),
        })

    total_questions = sum(len(q["questions"]) for q in quizzes.values())

    frappe.enqueue(
        "lms_plus.api.quiz_upload._create_quizzes_job",
        quizzes_data=list(quizzes.values()),
        queue="long",
        timeout=3600,
    )

    return {
        "status": "enqueued",
        "quizzes": len(quizzes),
        "questions": total_questions,
        "message": _("{0} quizzes with {1} questions queued for creation.").format(
            len(quizzes), total_questions
        ),
    }


def _create_quizzes_job(quizzes_data: list):
    """
    Background job — creates LMS Quiz and LMS Question records from Excel data.
    """
    created_quizzes, created_questions, skipped = 0, 0, 0

    for quiz_data in quizzes_data:
        slug = quiz_data["slug"]

        # Skip if quiz already exists
        if frappe.db.exists("LMS Quiz", slug):
            skipped += 1
            frappe.logger().info(f"LMS Plus quiz upload: quiz '{slug}' already exists — skipped.")
            continue

        # Create all questions first
        question_names = []
        for q in quiz_data["questions"]:
            try:
                q_fields = {
                    "doctype":   "LMS Question",
                    "question":  q["question"],
                    "type":      q["type"],
                    "multiple":  1 if q["multiple"] else 0,
                }

                if q["type"] == "User Input":
                    # User Input questions need at least one possible answer
                    # Use the question text as the placeholder answer
                    q_fields["possibility_1"] = q.get("possibility_1") or "See instructor for answer"
                else:
                    q_fields.update({
                        "option_1":     q["option_1"],
                        "is_correct_1": 1 if q["correct_1"] else 0,
                        "option_2":     q["option_2"],
                        "is_correct_2": 1 if q["correct_2"] else 0,
                        "option_3":     q["option_3"],
                        "is_correct_3": 1 if q["correct_3"] else 0,
                        "option_4":     q["option_4"],
                        "is_correct_4": 1 if q["correct_4"] else 0,
                    })

                q_doc = frappe.get_doc(q_fields)
                q_doc.insert(ignore_permissions=True)
                question_names.append({"name": q_doc.name, "marks": q["marks"], "type": q["type"]})
                created_questions += 1
            except Exception:
                frappe.log_error(
                    title=f"LMS Plus: failed to create question for quiz '{slug}'",
                    message=frappe.get_traceback(),
                )

        if not question_names:
            frappe.logger().warning(f"LMS Plus quiz upload: no questions created for '{slug}' — skipping quiz.")
            continue

        # Create the quiz
        try:
            total_marks = sum(q["marks"] for q in quiz_data["questions"])
            total_q = len(question_names)

            # limit_questions_to must be less than total questions
            # if not valid, set to 0 (no limit)
            limit = quiz_data["limit_questions_to"]
            if limit >= total_q:
                limit = 0

            quiz_doc = frappe.get_doc({
                "doctype":                 "LMS Quiz",
                "name":                    slug,
                "title":                   quiz_data["title"],
                "passing_percentage":      quiz_data["passing_percentage"],
                "max_attempts":            quiz_data["max_attempts"],
                "duration":                quiz_data["duration"],
                "shuffle_questions":       1 if quiz_data["shuffle_questions"] else 0,
                "limit_questions_to":      limit,
                "enable_negative_marking": 1 if quiz_data["enable_negative_marking"] else 0,
                "marks_to_cut":            quiz_data["marks_to_cut"],
                "total_marks":             total_marks,
            })

            for q in question_names:
                quiz_doc.append("questions", {
                    "question": q["name"],
                    "marks":    q["marks"],
                    "type":     q["type"],
                })

            quiz_doc.insert(ignore_permissions=True)
            created_quizzes += 1
            frappe.logger().info(f"LMS Plus: quiz '{slug}' created with {len(question_names)} questions.")
        except Exception:
            frappe.log_error(
                title=f"LMS Plus: failed to create quiz '{slug}'",
                message=frappe.get_traceback(),
            )

        frappe.db.commit()

    frappe.logger().info(
        f"LMS Plus quiz upload complete: "
        f"{created_quizzes} quizzes, {created_questions} questions created, {skipped} skipped."
    )
